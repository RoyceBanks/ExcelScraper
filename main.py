from unittest import skip
from openpyxl import load_workbook, Workbook                            #REQUIREMENTS 
from openpyxl.utils import get_column_letter
 

wb= load_workbook ('Pack.xlsx') # Or name of your file.xlsx             #Loads Excel file
ws = wb.active                                                          #Selects active Sheet

partials = 0                                                            #Variable
full = 0                                                                #Variable
#Pallet Count
for row in range (5, 900):                                              #Selects Rows 5-900
    for col in range(10, 11):                                           #Selects Column in the 10th spot over   
        char = get_column_letter(col)                                   #Identifies Column as J
        selected_box = ws[char + str(row)].value                        #Get whats entered in Selected box
        
        if selected_box == None:                                        #If the box is empty skip
            skip            
        else:
            pallets = str(selected_box)                                 #Math to add get Partial and Full pallet count
            if selected_box == "PARTIAL":
                partials = partials +1
            if selected_box == "FULL":
                full = full +1
            

ranch = 0                                                               #Variable
last = 0                                                                #Variable
#Last Pallet Received 
for row in range(5, 900):                                               #Selects Rows 5-900
    for col in range(7, 8):                                             #Selects Column in the 7th spot over
        char = get_column_letter(col)                                   #Identifies Column as G
        selected_box = ws[char + str(row)].value                        #Get whats entered in Selected box
        
        if selected_box == None :                                       #If the box is empty skip
            skip
        else:
            time = str(selected_box).split(" ")[-1]
            time = int(time.replace(":", ""))
        if time > last:
            last = time
        
        
for row in range(5, 900):                                             
    for col in range(1, 2):
        char = get_column_letter(col)
        selected_box = ws[char + str(row)].value
        if selected_box == None :
            skip
        else:
            ranch = selected_box
            
                    
print(f"""\n---------------------------------------------                 #Print statement
There are {full} Full and {partials} Partial Pallets 
Last pallet received at {last} by Ranch {ranch} 
---------------------------------------------- \n""")
